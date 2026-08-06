"""
Generador de exportación Excel (FR-096).

Usa openpyxl para producir un .xlsx con cabecera, filas y totales.
"""

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _to_float(value) -> float:
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def generate_excel(rows: list[dict]) -> bytes:
    """Genera el contenido del libro Excel (.xlsx) de los movimientos.

    Args:
        rows: Filas de movimientos (date, type, category, description, amount).

    Returns:
        Bytes del archivo XLSX.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimientos"

    headers = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    total = Decimal("0")
    for row in rows:
        amount = row["amount"]
        total += amount
        sheet.append(
            [
                row["date"],
                row["type"],
                row["category"],
                row["description"],
                _to_float(amount),
            ]
        )

    total_row = len(rows) + 2
    sheet.cell(row=total_row, column=4, value="Total").font = Font(bold=True)
    sheet.cell(row=total_row, column=5, value=_to_float(total)).font = Font(bold=True)

    for col_index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 20

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
